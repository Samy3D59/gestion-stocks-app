import pandas as pd
import streamlit as st
import os
from openpyxl import Workbook
from PIL import Image

# Configuration de la page
st.set_page_config(page_title="Gestion de Stocks", layout="wide")

# Définir les chemins
BASE_DIR = os.path.dirname(__file__)
FILE_PATH = os.path.join(BASE_DIR, "Gestion_Ventes_Stock.xlsx")
IMAGES_PATH = os.path.join(BASE_DIR, "IMAGES")
LOGO_PATH = os.path.join(IMAGES_PATH, "logo.png")

# Créer les dossiers nécessaires
if not os.path.exists(IMAGES_PATH):
    os.makedirs(IMAGES_PATH)

# Initialisation ou correction du fichier Excel
def initialize_excel():
    sheets_data = {
        "Stockage": [
            "Numéro de Commande", "Produit Vendu", "Quantité", "Prix Unitaire (Vendu)", "Frais Additionnels",
            "Frais de Livraison", "Transporteur", "Nom de l'Acheteur", "Commentaire", "Plateforme",
            "Date de Saisie", "Date d'Expédition", "Date de Livraison"
        ],
        "Stock": ["Nom du Produit", "SKU", "Quantité Initiale", "Quantité Restante", "Prix", "Image"],
        "Coffrets": ["Nom du Coffret", "Nom du Produit", "Quantité", "Image Coffret"]
    }

    if os.path.exists(FILE_PATH):
        try:
            pd.ExcelFile(FILE_PATH)
        except Exception:
            os.remove(FILE_PATH)
            st.warning("Fichier Excel corrompu détecté. Un nouveau fichier sera créé.")

    if not os.path.exists(FILE_PATH):
        wb = Workbook()
        for sheet_name, headers in sheets_data.items():
            ws = wb.create_sheet(sheet_name)
            ws.append(headers)
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        wb.save(FILE_PATH)
    else:
        with pd.ExcelWriter(FILE_PATH, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            for sheet_name, headers in sheets_data.items():
                try:
                    df = pd.read_excel(FILE_PATH, sheet_name=sheet_name)
                    for header in headers:
                        if header not in df.columns:
                            df[header] = None
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                except ValueError:
                    empty_df = pd.DataFrame(columns=headers)
                    empty_df.to_excel(writer, sheet_name=sheet_name, index=False)

initialize_excel()

# Charger les données d'une feuille Excel
def load_sheet(sheet_name):
    try:
        return pd.read_excel(FILE_PATH, sheet_name=sheet_name)
    except Exception as e:
        st.error(f"Erreur lors du chargement de la feuille {sheet_name}: {e}")
        return pd.DataFrame()

# Sauvegarder les données dans une feuille Excel
def save_to_excel(df, sheet_name):
    with pd.ExcelWriter(FILE_PATH, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

# Chargement des données
stockage_df = load_sheet("Stockage")
stock_df = load_sheet("Stock")
coffret_df = load_sheet("Coffrets")

# Ajouter les colonnes calculées si elles n'existent pas
if "Prix Total Vente (€)" not in stockage_df.columns:
    stockage_df["Prix Total Vente (€)"] = (
        stockage_df["Quantité"] * stockage_df["Prix Unitaire (Vendu)"]
        - stockage_df["Frais Additionnels"]
        - stockage_df["Frais de Livraison"]
    )

if "Coût Total Revient (€)" not in stockage_df.columns:
    stockage_df["Coût Total Revient (€)"] = stockage_df["Quantité"] * stockage_df["Prix Unitaire (Vendu)"] * 0.6  # Exemple : 60% de marge brute

if "Profit Net (€)" not in stockage_df.columns:
    stockage_df["Profit Net (€)"] = stockage_df["Prix Total Vente (€)"] - stockage_df["Coût Total Revient (€)"]

if "Marge Nette (%)" not in stockage_df.columns:
    stockage_df["Marge Nette (%)"] = (
        (stockage_df["Profit Net (€)"] / stockage_df["Prix Total Vente (€)"]) * 100
    ).fillna(0)

# Afficher le logo en haut
if os.path.exists(LOGO_PATH):
    st.image(LOGO_PATH, use_column_width=False, width=150)

# Titre de l'application
st.title("📦 Gestion de Stocks et des Ventes")
st.markdown("### Une interface intuitive pour gérer vos produits, coffrets et ventes")

# Sidebar : Navigation
st.sidebar.header("Navigation")
page = st.sidebar.radio("Menu", ["Tableau de Bord", "Gestion du Stock", "Gestion des Coffrets", "Gestion des Ventes"])

# Tableau de Bord
if page == "Tableau de Bord":
    st.header("📊 Tableau de Bord")
    if not stock_df.empty:
        for _, row in stock_df.iterrows():
            col1, col2 = st.columns([1, 3])
            with col1:
                if row["Image"] and os.path.exists(row["Image"]):
                    st.image(row["Image"], width=100)
                else:
                    st.warning(f"L'image pour le produit '{row['Nom du Produit']}' est introuvable.")
            with col2:
                st.markdown(f"**Produit :** {row['Nom du Produit']} (SKU: {row['SKU']})")
                st.markdown(f"**Quantité Restante :** {row['Quantité Restante']}")
                st.markdown(f"**Prix :** {row['Prix']} €")
    else:
        st.info("Le stock est vide. Ajoutez des produits dans la section 'Gestion du Stock'.")

# Gestion du Stock
elif page == "Gestion du Stock":
    st.header("🛠️ Gestion du Stock")
    with st.form("add_product_form"):
        produit = st.text_input("Nom du Produit")
        sku = st.text_input("SKU")
        quantite_initiale = st.number_input("Quantité Initiale", min_value=0)
        prix = st.number_input("Prix Unitaire (€)", min_value=0.0, step=0.01)
        image = st.file_uploader("Téléchargez une image du produit", type=["png", "jpg", "jpeg"])
        submit = st.form_submit_button("Ajouter au Stock")

        if submit:
            if produit and sku and image:
                image_path = os.path.join(IMAGES_PATH, f"{sku}.jpg")
                with open(image_path, "wb") as f:
                    f.write(image.getbuffer())
                new_product = pd.DataFrame({
                    "Nom du Produit": [produit],
                    "SKU": [sku],
                    "Quantité Initiale": [quantite_initiale],
                    "Quantité Restante": [quantite_initiale],
                    "Prix": [prix],
                    "Image": [image_path]
                })
                stock_df = pd.concat([stock_df, new_product], ignore_index=True)
                save_to_excel(stock_df, "Stock")
                st.success("Produit ajouté avec succès, image enregistrée !")
            else:
                st.error("Veuillez remplir tous les champs et télécharger une image.")

    st.subheader("Stock actuel")
    if not stock_df.empty:
        st.dataframe(stock_df.drop(columns=["Image"]))
    else:
        st.info("Aucun produit dans le stock pour le moment.")

# Gestion des Coffrets
elif page == "Gestion des Coffrets":
    st.header("🎁 Gestion des Coffrets")
    coffret_nom = st.text_input("Nom du Coffret")
    image_coffret = st.file_uploader("Téléchargez une image pour le coffret", type=["png", "jpg", "jpeg"])

    if not stock_df.empty:
        produits_selectionnes = st.multiselect(
            "Produits disponibles",
            stock_df["Nom du Produit"].unique()
        )
        quantites = {}
        for produit in produits_selectionnes:
            quantites[produit] = st.number_input(f"Quantité pour {produit}", min_value=1, value=1)

        if st.button("Créer le Coffret"):
            if coffret_nom and produits_selectionnes:
                image_path = None
                if image_coffret:
                    image_path = os.path.join(IMAGES_PATH, f"{coffret_nom}_coffret.jpg")
                    with open(image_path, "wb") as f:
                        f.write(image_coffret.getbuffer())

                coffret_data = []
                for produit, qte in quantites.items():
                    coffret_data.append({
                        "Nom du Coffret": coffret_nom,
                        "Nom du Produit": produit,
                        "Quantité": qte,
                        "Image Coffret": image_path
                    })
                coffret_df = pd.concat([coffret_df, pd.DataFrame(coffret_data)], ignore_index=True)
                save_to_excel(coffret_df, "Coffrets")
                st.success(f"Coffret '{coffret_nom}' créé avec succès !")
            else:
                st.error("Veuillez fournir un nom de coffret et sélectionner au moins un produit.")
    else:
        st.warning("Aucun produit disponible pour créer un coffret.")

    st.subheader("📦 Liste des Coffrets")
    if not coffret_df.empty:
        coffrets = coffret_df["Nom du Coffret"].unique()
        for coffret in coffrets:
            st.markdown(f"### Coffret : **{coffret}**")
            coffret_image_path = coffret_df.loc[coffret_df["Nom du Coffret"] == coffret, "Image Coffret"].iloc[0]
            if coffret_image_path and os.path.exists(coffret_image_path):
                st.image(coffret_image_path, caption=f"Coffret : {coffret}", width=200)
            else:
                st.warning("Pas d'image associée à ce coffret.")

            st.markdown("#### Produits inclus dans le coffret :")
            produits = coffret_df[coffret_df["Nom du Coffret"] == coffret]
            for _, row in produits.iterrows():
                produit_nom = row["Nom du Produit"]
                quantite = row["Quantité"]
                produit_image_path = stock_df.loc[stock_df["Nom du Produit"] == produit_nom, "Image"].iloc[0]

                col1, col2 = st.columns([1, 3])
                with col1:
                    if produit_image_path and os.path.exists(produit_image_path):
                        st.image(produit_image_path, caption=produit_nom, width=100)
                    else:
                        st.warning(f"Pas d'image pour le produit : {produit_nom}")
                with col2:
                    st.markdown(f"- **Produit :** {produit_nom}")
                    st.markdown(f"  - Quantité : {quantite}")

# Gestion des Ventes
elif page == "Gestion des Ventes":
    st.header("💰 Gestion des Ventes")

    # Type de vente (Produit individuel ou Coffret)
    vente_type = st.radio("Type de Vente", ["Produit individuel", "Coffret"])

    commande_num = st.text_input("Numéro de Commande")
    if vente_type == "Produit individuel":
        st.subheader("Vente d'un produit individuel")
        produit_vendu = st.selectbox("Produit Vendu", stock_df["Nom du Produit"].unique())
        quantite_vendue = st.number_input("Quantité Vendue", min_value=1, value=1)
    elif vente_type == "Coffret":
        st.subheader("Vente d'un coffret")
        produit_vendu = st.selectbox("Coffret Vendu", coffret_df["Nom du Coffret"].unique())
        quantite_vendue = st.number_input("Quantité de Coffrets Vendus", min_value=1, value=1)

    # Champs supplémentaires
    prix_unitaire = st.number_input("Prix Unitaire (Vendu)", min_value=0.0, step=0.01)
    frais_additionnels = st.number_input("Frais Additionnels", min_value=0.0, step=0.01)
    frais_livraison = st.number_input("Frais de Livraison", min_value=0.0, step=0.01)
    transporteur = st.selectbox("Transporteur", ["Colissimo", "UPS", "Mondial Relay", "La Poste", "Autre"])
    nom_acheteur = st.text_input("Nom de l'Acheteur")
    commentaire = st.text_area("Commentaire (facultatif)")
    plateforme = st.selectbox("Plateforme", ["Amazon", "Vinted", "Réseaux", "Autre"])
    date_vente = st.date_input("Date de Vente")
    date_expedition = st.date_input("Date d'Expédition")
    date_livraison = st.date_input("Date de Livraison")

    # Calcul automatique des coûts
    prix_total_vente = (quantite_vendue * prix_unitaire) - frais_additionnels - frais_livraison
    cout_unitaire = stock_df.loc[stock_df["Nom du Produit"] == produit_vendu, "Prix"].values[0] if vente_type == "Produit individuel" else 0
    cout_total_revient = quantite_vendue * cout_unitaire
    profit_net = prix_total_vente - cout_total_revient
    marge_nette = (profit_net / prix_total_vente * 100) if prix_total_vente > 0 else 0

    # Affichage des calculs
    st.markdown("### Résumé des Calculs")
    st.markdown(f"- **Prix de Vente Total :** {prix_total_vente:.2f} €")
    st.markdown(f"- **Coût de Revient Total :** {cout_total_revient:.2f} €")
    st.markdown(f"- **Profit Net :** {profit_net:.2f} €")
    st.markdown(f"- **Marge Nette :** {marge_nette:.2f} %")

    if st.button("Enregistrer la Vente"):
        # Gestion des stocks
        if vente_type == "Produit individuel":
            current_quantity = stock_df.loc[stock_df["Nom du Produit"] == produit_vendu, "Quantité Restante"].values[0]
            if quantite_vendue > current_quantity:
                st.error("La quantité vendue dépasse la quantité restante en stock !")
                st.stop()
            stock_df.loc[stock_df["Nom du Produit"] == produit_vendu, "Quantité Restante"] -= quantite_vendue
        elif vente_type == "Coffret":
            produits_coffret = coffret_df[coffret_df["Nom du Coffret"] == produit_vendu]
            stock_insuffisant = []
            for _, row in produits_coffret.iterrows():
                produit_nom = row["Nom du Produit"]
                produit_qte = row["Quantité"] * quantite_vendue
                stock_dispo = stock_df.loc[stock_df["Nom du Produit"] == produit_nom, "Quantité Restante"].values[0]
                if produit_qte > stock_dispo:
                    stock_insuffisant.append(produit_nom)
            if stock_insuffisant:
                st.error(f"Stock insuffisant pour les produits suivants : {', '.join(stock_insuffisant)}")
                st.stop()
            for _, row in produits_coffret.iterrows():
                produit_nom = row["Nom du Produit"]
                produit_qte = row["Quantité"] * quantite_vendue
                stock_df.loc[stock_df["Nom du Produit"] == produit_nom, "Quantité Restante"] -= produit_qte

        # Enregistrement de la vente
        new_sale = pd.DataFrame({
            "Numéro de Commande": [commande_num],
            "Produit Vendu": [produit_vendu],
            "Quantité": [quantite_vendue],
            "Prix Unitaire (Vendu)": [prix_unitaire],
            "Frais Additionnels": [frais_additionnels],
            "Frais de Livraison": [frais_livraison],
            "Transporteur": [transporteur],
            "Nom de l'Acheteur": [nom_acheteur],
            "Commentaire": [commentaire],
            "Plateforme": [plateforme],
            "Prix Total Vente (€)": [prix_total_vente],
            "Coût Total Revient (€)": [cout_total_revient],
            "Profit Net (€)": [profit_net],
            "Marge Nette (%)": [marge_nette],
            "Date de Saisie": [pd.Timestamp(date_vente)],
            "Date d'Expédition": [pd.Timestamp(date_expedition)],
            "Date de Livraison": [pd.Timestamp(date_livraison)],
        })
        stockage_df = pd.concat([stockage_df, new_sale], ignore_index=True)

        # Sauvegarde
        save_to_excel(stock_df, "Stock")
        save_to_excel(stockage_df, "Stockage")

        st.success(f"Vente enregistrée avec succès : {quantite_vendue} unité(s) de '{produit_vendu}'")

    # Historique des Ventes
    st.subheader("📜 Historique des Ventes")
    if not stockage_df.empty:
        st.markdown("### Filtres")
        filter_transporteur = st.multiselect("Transporteur", stockage_df["Transporteur"].unique())
        filter_plateforme = st.multiselect("Plateforme", stockage_df["Plateforme"].unique())
        filter_date_debut = st.date_input("Date Début", value=pd.to_datetime("2024-01-01").date())
        filter_date_fin = st.date_input("Date Fin", value=pd.to_datetime("2024-12-31").date())

        # Convertir les dates Streamlit en Timestamp pour alignement
        filter_date_debut = pd.Timestamp(filter_date_debut)
        filter_date_fin = pd.Timestamp(filter_date_fin)

        # Application des filtres
        filtered_df = stockage_df.copy()
        if filter_transporteur:
            filtered_df = filtered_df[filtered_df["Transporteur"].isin(filter_transporteur)]
        if filter_plateforme:
            filtered_df = filtered_df[filtered_df["Plateforme"].isin(filter_plateforme)]
        filtered_df = filtered_df[
            (filtered_df["Date de Saisie"] >= filter_date_debut) &
            (filtered_df["Date de Saisie"] <= filter_date_fin)
        ]

        # Affichage du tableau épuré
        st.dataframe(filtered_df[[
            "Numéro de Commande", "Produit Vendu", "Quantité", "Prix Unitaire (Vendu)",
            "Prix Total Vente (€)", "Coût Total Revient (€)", "Profit Net (€)", "Marge Nette (%)",
            "Transporteur", "Nom de l'Acheteur", "Plateforme", "Date de Saisie"
        ]])

        # Téléchargement des ventes filtrées
        st.markdown("### Télécharger les Ventes Filtrées")
        csv = filtered_df.to_csv(index=False).encode("utf-8")
        st.download_button("Télécharger en CSV", data=csv, file_name="ventes_filtrees.csv", mime="text/csv")
    else:
        st.info("Aucune vente enregistrée pour le moment.")
